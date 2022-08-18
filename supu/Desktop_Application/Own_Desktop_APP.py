# Tkinterでデスクトップアプリを作成
# openpyxlやtwitterなどと組み合わせて、自分のアプリを作れる
import tkinter
from  pathlib import  Path
from tkinter import  filedialog # データ保存用
import  openpyxl
import time

class Application(tkinter.Frame):   # フレームオブジェクト作成、アプリの中身を書く
    def __init__(self, root=None):
        super().__init__(root, width=380, height=280, borderwidth=1, relief='groove')   # 横幅、高さ、境界線の太さ、境界線の種類
        self.root = root    # 引数として保存、クラス内で容易に扱える
        self.pack() # 位置設定
        self.pack_propagate(False)  # サイズ調整
        self.create_widgets()

    def create_widgets(self):   # ここに部品を追加していく
        # 閉じるボタン
        quit_btn = tkinter.Button(self) # ボタンクラスオブジェクト生成
        quit_btn['text'] = '閉じる'    # 表示させる文字
        quit_btn['command'] = self.root.destroy # ボタンに押された時の処理
        quit_btn.pack(side='bottom')    # 位置設定：下

        # テキストボックス
        self.text_box = tkinter.Entry(self)
        self.text_box['width'] = 10
        self.text_box.pack()

        # 実行ボタン
        submit_btn = tkinter.Button(self)
        submit_btn['text'] = '実行'
        submit_btn['command'] = self.save_data
        submit_btn.pack()

        # メッセージ出力
        self.message = tkinter.Message(self)
        self.message.pack()

    # 保存するタイプのアプリ：データ永続化
    # Tkinter x Excel
    def save_data(self):
        self.message['text'] = '最初の一万件しか保存できません'
        text = self.text_box.get()
        file_name = tkinter.filedialog.askopenfilename(initialfile=(Path.cwd() / 'data.xlsx'))
        wb = openpyxl.load_workbook(file_name)
        ws = wb.worksheets[0]
        if ws['A1'].value >= 10000:
            ws['A1'].value = 0  # 最初のマスから書き込み、そのデータが上書きされる\A[0 + 1] = A[1]
        mem = 1 if not ws['A1'].value else ws['A1'].value + 1   # 回数記録\value付いてないと計算できん
        ws['A1'] = mem    # 回数書き込み
        ws[f'B{mem}'] = text    # 改行
        wb.save(file_name)
        time.sleep(0.7)  # 0.7秒表示させる
        self.message['text'] = '保存完了'

root = tkinter.Tk() # アプリの土台
root.title('Kaniのアプリ')  # タイトル
root.geometry('400x300')    # ピクセル
app = Application(root=root)    # フレーム
app.mainloop() # 起動
