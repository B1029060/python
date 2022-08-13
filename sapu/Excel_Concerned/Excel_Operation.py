# PythonでのExcel操作について

# OpenPyXL
# pip install openpyxl
import random

from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook() # Excelファイル新規作成
ws = wb.active  # シートの作成
ws['A1'] = 10   # マスに値を指定

cel = ws['A1']
cel.font = Font(size=12, bold=True, color='FF0000') # フォント変更
wb.save('output.xlsx')  # ファイルセーブ

from openpyxl import load_workbook  # 既存Excelファイル読み込み
wb = load_workbook(filename='output.xlsx')
sheet_names = wb.sheetnames[0] # 最初のシートの名前を取得
print(sheet_names)
ws1 = wb[sheet_names]
x = ws1['A1'].value # 当マスの値を取得
ws2 = wb.create_sheet('new sheet')  # 新シートを作る
ws2['A1'] = x
wb.save('output.xlsx')

# Excelで良く使う操作について
# セル結合
wb = load_workbook(filename='output.xlsx')
sheet_names = wb.sheetnames[0]
ws1 = wb[sheet_names]
ws1.merge_cells('A5:B7')    # 範囲
ws1['A5'] = 'Kani'  # 一番左上のセルに値指定
# 罫線を引く
from openpyxl.styles.borders import Border, Side
s = Side(style='thin', color='FF0000')  # thin, double, thick/線を引く
b = Border(left=s, right=s, top=s, bottom=s)    # 周りの線の設定
cell = ws1['A2']
cell.border = b # ボーダー設置(罫線)
# 列、行の操作
ws2 = wb[wb.sheetnames[1]]

x_list = [] # A1からC10までの値を埋め込む
for x in range(1, 11):
    x_list.append(f'A{x}')
    x_list.append(f'B{x}')
    x_list.append(f'C{x}')
for x in x_list:
    ws2[x] = x  # 終了

ws2.insert_cols(2, 3)   # 2列目から3列を挿入
ws2.insert_rows(4, 5)   # 4行目から5行を挿入
ws2.delete_rows(2)  # 2行目を削除
# グラフの描画
from openpyxl.chart import  BarChart, Reference # BarChart, LineChart, PieChart

y_list = [] # B2からC13までの値を埋め込む
for y in range(2, 14):
    y_list.append(f'B{y}')
    y_list.append(f'{y - 1}月')
    y_list.append(f'C{y}')
    y_list.append(random.randint(8000, 100000)) # 適当に値を書き込む
ws3 = wb.create_sheet('new new sheet')
for y in range(0, len(y_list), 4):
    ws3[y_list[y]] = y_list[y + 1]
    ws3[y_list[y + 2]] = y_list[y + 3]  #終了

title =Reference(ws3, min_col=2, min_row=2, max_col=2, max_row=13)  # 1月から12月までのタイトルを取得
data = Reference(ws3, min_col=3, min_row=2, max_col=3, max_row=13)  # データを取得
chart = BarChart()  # グラフの型：棒グラフ
chart.add_data(data)# データを入れる/y軸
chart.set_categories(title) # x軸
ws3.add_chart(chart, 'D2')  # どのセルにチャート挿入かを選択
wb.save('output.xlsx')
